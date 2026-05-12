from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
import csv
import io
import re
import ssl
import unicodedata
from typing import Any, Iterable
from urllib.parse import urlencode
from urllib.request import urlopen

from openpyxl import load_workbook


STATUS_WEIGHTS = {
    "vigor": 4,
    "solicitada": 3,
    "suspendida": 2,
    "anulada": 1,
}

MONTH_NAMES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

MONTH_LABELS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

INSURANCE_SYNONYMS = {
    "auto": "AUTOS",
    "autos": "AUTOS",
    "coche": "AUTOS",
    "hogar": "MULTIRIESGO HOGAR",
    "casa": "MULTIRIESGO HOGAR",
    "salud": "ASISTENCIA SANITARIA",
    "sanitario": "ASISTENCIA SANITARIA",
    "sanitaria": "ASISTENCIA SANITARIA",
    "comercio": "MULTIRIESGO COMERCIO",
    "comunidad": "MULTIRIESGO COMUNIDADES",
    "vida": "VIDA INDIVIDUAL",
    "barco": "EMBARCACIONES",
    "embarcacion": "EMBARCACIONES",
    "accidente": "ACCIDENTES",
    "deceso": "DECESOS",
}

POSTAL_PREFIX_COORDS = {
    "01": (-2.68, 42.85), "02": (-1.86, 38.99), "03": (-0.49, 38.35), "04": (-2.46, 36.84),
    "05": (-4.70, 40.66), "06": (-6.97, 38.88), "07": (2.65, 39.57), "08": (2.17, 41.38),
    "09": (-3.70, 42.34), "10": (-6.37, 39.47), "11": (-6.29, 36.53), "12": (-0.04, 39.99),
    "13": (-3.93, 38.99), "14": (-4.78, 37.88), "15": (-8.41, 43.36), "16": (-2.14, 40.07),
    "17": (2.82, 41.98), "18": (-3.60, 37.18), "19": (-3.16, 40.63), "20": (-1.98, 43.32),
    "21": (-6.94, 37.26), "22": (-0.41, 42.14), "23": (-3.79, 37.77), "24": (-5.57, 42.60),
    "25": (0.62, 41.62), "26": (-2.45, 42.47), "27": (-7.56, 43.01), "28": (-3.70, 40.42),
    "29": (-4.42, 36.72), "30": (-1.13, 37.99), "31": (-1.64, 42.82), "32": (-7.86, 42.34),
    "33": (-5.85, 43.36), "34": (-4.53, 42.01), "35": (-15.43, 28.12), "36": (-8.64, 42.43),
    "37": (-5.66, 40.97), "38": (-16.25, 28.46), "39": (-3.81, 43.46), "40": (-4.11, 40.95),
    "41": (-5.99, 37.39), "42": (-2.47, 41.76), "43": (1.25, 41.12), "44": (-1.11, 40.34),
    "45": (-4.02, 39.86), "46": (-0.38, 39.47), "47": (-4.72, 41.65), "48": (-2.94, 43.26),
    "49": (-5.74, 41.50), "50": (-0.88, 41.65), "51": (-5.31, 35.89), "52": (-2.94, 35.29),
}

LOCALITY_COORDS = {
    "corralejo": (-13.867, 28.731),
    "playa corralejo": (-13.867, 28.731),
    "geafond": (-13.875, 28.719),
    "lajares": (-13.938, 28.680),
    "los lajares": (-13.938, 28.680),
    "villaverde": (-13.914, 28.629),
    "cotillo": (-14.008, 28.682),
    "el cotillo": (-14.008, 28.682),
    "la oliva": (-13.928, 28.611),
    "puerto del rosario": (-13.862, 28.500),
    "costa calma": (-14.225, 28.161),
    "puerto rico": (-15.710, 27.789),
    "las palmas de gran canaria": (-15.413, 28.124),
    "las palmas de gran canari": (-15.413, 28.124),
    "las palmas": (-15.413, 28.124),
    "telde": (-15.416, 27.995),
    "arguineguin": (-15.682, 27.760),
    "maspalomas": (-15.586, 27.761),
    "playa del ingles": (-15.575, 27.756),
    "mogan": (-15.725, 27.883),
    "san bartolome": (-15.573, 27.924),
    "vecindario": (-15.445, 27.846),
    "ingenio": (-15.443, 27.918),
    "arrecife": (-13.548, 28.963),
    "playa blanca": (-13.828, 28.864),
    "tias": (-13.651, 28.953),
    "puerto del carmen": (-13.666, 28.923),
}


@dataclass(frozen=True)
class PolicyRecord:
    policyNumber: str
    clientId: str
    clientName: str
    email: str
    status: str
    officeId: str
    officeName: str
    insurerName: str
    insuranceType: str
    postalCode: str
    locality: str
    premiumNet: float
    expirationDate: date | None
    contractDate: date | None
    cancellationDate: date | None
    clientDropDate: date | None
    clientDropReason: str
    isActive: bool

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        for key in ("expirationDate", "contractDate", "cancellationDate", "clientDropDate"):
            payload[key] = payload[key].isoformat() if payload[key] else None
        return payload


@dataclass(frozen=True)
class ClientSummary:
    clientId: str
    clientName: str
    email: str
    activePolicyCount: int
    totalPolicyCount: int
    isActiveClient: bool
    isFormerClient: bool
    lastPolicyType: str
    lastPolicyNumber: str
    lastExpirationDate: date | None
    lastContractDate: date | None
    clientDropDate: date | None
    clientDropReason: str
    officeIds: list[str]

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        for key in ("lastExpirationDate", "lastContractDate", "clientDropDate"):
            payload[key] = payload[key].isoformat() if payload[key] else None
        return payload


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def normalize_email(value: Any) -> str:
    return normalize_text(value).lower()


def strip_accents(value: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFD", value) if unicodedata.category(char) != "Mn"
    )


def normalize_for_match(value: str) -> str:
    value = strip_accents(value.lower())
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def parse_excel_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_mixed_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    if not text:
        return 0.0
    cleaned = re.sub(r"[^0-9,.\-]", "", text)
    if not cleaned or cleaned == "-":
        return 0.0

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def normalize_postal_code(value: Any) -> str:
    digits = re.sub(r"\D", "", normalize_text(value))
    if not digits:
        return "Sin CP"
    return digits.zfill(5)[:5]


def postal_code_point(postal_code: str) -> dict[str, float | str] | None:
    code = normalize_postal_code(postal_code)
    if code == "Sin CP":
        return None
    coords = POSTAL_PREFIX_COORDS.get(code[:2])
    if not coords:
        return None
    lon, lat = coords
    min_lon, max_lon = -18.5, 4.5
    min_lat, max_lat = 27.2, 43.9
    x = ((lon - min_lon) / (max_lon - min_lon)) * 100
    y = (1 - ((lat - min_lat) / (max_lat - min_lat))) * 100
    local = int(code[2:] or "0")
    x += ((local % 9) - 4) * 0.72
    y += (((local // 9) % 9) - 4) * 0.72
    x = min(max(x, 3), 97)
    y = min(max(y, 3), 97)
    return {"cp": code, "lon": lon, "lat": lat, "x": round(x, 2), "y": round(y, 2)}


def locality_coordinates(locality: str) -> tuple[float, float] | None:
    normalized = normalize_for_match(locality)
    for key, coords in LOCALITY_COORDS.items():
        if key in normalized:
            return coords
    return None


def dedupe_headers(headers: Iterable[Any]) -> list[str]:
    counts: dict[str, int] = {}
    output: list[str] = []
    for index, header in enumerate(headers):
        base = normalize_text(header) or f"Unnamed: {index}"
        count = counts.get(base, 0)
        if count:
            output.append(f"{base}.{count}")
        else:
            output.append(base)
        counts[base] = count + 1
    return output


def build_client_key(client_alias: str, nif: str, client_name: str, policy_number: str) -> str:
    if client_alias:
        return client_alias
    if nif:
        return nif
    if client_name:
        return normalize_for_match(client_name)
    return f"unknown-{policy_number}"


def record_quality(record: PolicyRecord) -> tuple[int, int, int, int, float]:
    filled = sum(
        int(bool(value))
        for value in (
            record.clientName,
            record.email,
            record.officeId,
            record.insuranceType,
            record.clientDropReason,
        )
    )
    date_score = sum(int(value is not None) for value in (record.expirationDate, record.contractDate, record.clientDropDate))
    return (
        STATUS_WEIGHTS.get(record.status.lower(), 0),
        filled,
        date_score,
        int(record.premiumNet != 0),
        record.premiumNet,
    )


class DashboardDataset:
    def __init__(self, policies: list[PolicyRecord], source_path: str):
        self.policies = policies
        self.source_path = source_path
        self.generated_at = datetime.now().isoformat(timespec="seconds")
        self.office_options = sorted({policy.officeId for policy in policies}, key=lambda item: (len(item), item))
        self.office_labels = {
            office_id: Counter(
                policy.officeName for policy in policies if policy.officeId == office_id and policy.officeName
            ).most_common(1)[0][0]
            for office_id in self.office_options
            if any(policy.officeName for policy in policies if policy.officeId == office_id)
        }
        self.insurer_options = sorted({policy.insurerName for policy in policies if policy.insurerName})
        self.insurance_options = sorted({policy.insuranceType for policy in policies})
        self.status_options = sorted({policy.status for policy in policies if policy.status})
        self.expiration_month_options = sorted(
            {
                policy.expirationDate.strftime("%Y-%m")
                for policy in policies
                if policy.expirationDate
            },
            reverse=True,
        )
        self._insurance_lookup = {normalize_for_match(item): item for item in self.insurance_options}

    @classmethod
    def from_excel(cls, excel_path: str) -> "DashboardDataset":
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(min_row=2, values_only=True)
        headers = dedupe_headers(next(rows))
        dataset = cls.from_rows(headers, rows, excel_path)
        workbook.close()
        return dataset

    @classmethod
    def from_google_sheet(cls, spreadsheet_id: str, gid: str = "674567918") -> "DashboardDataset":
        params = urlencode({"format": "csv", "gid": gid})
        url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?{params}"
        with urlopen(url, timeout=60, context=ssl._create_unverified_context()) as response:  # noqa: S323
            content = response.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        headers = dedupe_headers(next(reader))
        return cls.from_rows(headers, reader, f"Google Sheet Data PMP ({spreadsheet_id})")

    @classmethod
    def from_rows(cls, headers: list[str], rows: Iterable[Iterable[Any]], source_path: str) -> "DashboardDataset":

        deduped: dict[tuple[str, str, str], PolicyRecord] = {}
        for row in rows:
            if not row:
                continue
            row = list(row)
            values = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
            if not any(value not in (None, "") for value in values.values()):
                continue

            policy_number = normalize_text(values.get("Póliza"))
            client_alias = normalize_text(values.get("Cliente.Alias"))
            nif = normalize_text(values.get("NIF"))
            client_name = normalize_text(values.get("NombreCompleto2")) or normalize_text(values.get("NombreCompleto2.1"))
            if not policy_number and not client_alias and not client_name:
                continue

            expiration_date = parse_excel_date(values.get("Vencimiento"))
            client_key = build_client_key(client_alias, nif, client_name, policy_number)
            record = PolicyRecord(
                policyNumber=policy_number,
                clientId=client_key,
                clientName=client_name or client_key,
                email=normalize_email(values.get("Cliente.Contacto.Email") or values.get("Correo electrónico")),
                status=normalize_text(values.get("Estado")) or "Desconocido",
                officeId=normalize_text(values.get("OficinaComercialId")) or "Sin oficina",
                officeName=normalize_text(values.get("OficinaComercial.Nombre"))
                or normalize_text(values.get("Alias oficina"))
                or "Sin oficina",
                insurerName=normalize_text(values.get("Compania.Nombre"))
                or normalize_text(values.get("Compañia.Nombre"))
                or "Sin aseguradora",
                insuranceType=normalize_text(values.get("Producto.Ramo.Nombre"))
                or normalize_text(values.get("Alias ramo"))
                or "Sin tipo",
                postalCode=normalize_postal_code(values.get("Cliente.Direccion.CP")),
                locality=normalize_text(values.get("Cliente.Direccion.Localidad")) or "Sin localidad",
                premiumNet=parse_mixed_number(values.get("Prima neta")),
                expirationDate=expiration_date,
                contractDate=parse_excel_date(values.get("Fecha contratación")),
                cancellationDate=parse_excel_date(values.get("Fecha anulación")),
                clientDropDate=parse_excel_date(values.get("Cliente.FechaBaja")),
                clientDropReason=normalize_text(values.get("Cliente.MotivoBaja")),
                isActive=normalize_text(values.get("Estado")).lower() == "vigor",
            )

            dedupe_key = (
                record.policyNumber or f"sin-poliza-{record.clientId}",
                record.clientId,
                record.expirationDate.isoformat() if record.expirationDate else "",
            )
            existing = deduped.get(dedupe_key)
            if existing is None or record_quality(record) > record_quality(existing):
                deduped[dedupe_key] = record

        return cls(sorted(deduped.values(), key=lambda item: (item.clientName, item.policyNumber)), source_path)

    def office_label(self, office_id: str) -> str:
        return self.office_labels.get(office_id, f"Oficina {office_id}")

    def options_payload(self) -> dict[str, Any]:
        return {
            "offices": [{"id": office_id, "label": self.office_label(office_id)} for office_id in self.office_options],
            "statuses": self.status_options,
            "insurers": self.insurer_options,
            "insuranceTypes": self.insurance_options,
            "expirationMonths": self.expiration_month_options,
        }

    def apply_filters(
        self,
        office_ids: Iterable[str] | None = None,
        statuses: Iterable[str] | None = None,
        insurers: Iterable[str] | None = None,
        insurance_types: Iterable[str] | None = None,
        expiration_from: date | None = None,
        expiration_to: date | None = None,
    ) -> list[PolicyRecord]:
        office_set = {normalize_text(item) for item in (office_ids or []) if normalize_text(item)}
        status_set = {normalize_text(item) for item in (statuses or []) if normalize_text(item)}
        insurer_set = {normalize_text(item) for item in (insurers or []) if normalize_text(item)}
        type_set = {normalize_text(item) for item in (insurance_types or []) if normalize_text(item)}
        filtered: list[PolicyRecord] = []
        for policy in self.policies:
            if office_set and policy.officeId not in office_set:
                continue
            if status_set and policy.status not in status_set:
                continue
            if insurer_set and policy.insurerName not in insurer_set:
                continue
            if type_set and policy.insuranceType not in type_set:
                continue
            if expiration_from and (policy.expirationDate is None or policy.expirationDate < expiration_from):
                continue
            if expiration_to and (policy.expirationDate is None or policy.expirationDate > expiration_to):
                continue
            filtered.append(policy)
        return filtered

    def build_client_summaries(self, policies: list[PolicyRecord]) -> list[ClientSummary]:
        grouped: dict[str, list[PolicyRecord]] = defaultdict(list)
        for policy in policies:
            grouped[policy.clientId].append(policy)

        summaries: list[ClientSummary] = []
        for client_id, client_policies in grouped.items():
            active_policies = {policy.policyNumber for policy in client_policies if policy.isActive}
            all_policies = {policy.policyNumber for policy in client_policies if policy.policyNumber}
            emails = [policy.email for policy in client_policies if policy.email]
            email = Counter(emails).most_common(1)[0][0] if emails else ""
            last_policy = max(
                client_policies,
                key=lambda policy: (
                    policy.expirationDate or date.min,
                    policy.contractDate or date.min,
                    policy.policyNumber,
                ),
            )
            drop_date = max((policy.clientDropDate for policy in client_policies if policy.clientDropDate), default=None)
            drop_reason = next((policy.clientDropReason for policy in client_policies if policy.clientDropReason), "")
            is_active = bool(active_policies)
            summaries.append(
                ClientSummary(
                    clientId=client_id,
                    clientName=next((policy.clientName for policy in client_policies if policy.clientName), client_id),
                    email=email,
                    activePolicyCount=len(active_policies),
                    totalPolicyCount=len(all_policies),
                    isActiveClient=is_active,
                    isFormerClient=(not is_active) and bool(drop_date or drop_reason),
                    lastPolicyType=last_policy.insuranceType,
                    lastPolicyNumber=last_policy.policyNumber,
                    lastExpirationDate=last_policy.expirationDate,
                    lastContractDate=last_policy.contractDate,
                    clientDropDate=drop_date,
                    clientDropReason=drop_reason,
                    officeIds=sorted({policy.officeId for policy in client_policies}, key=lambda item: (len(item), item)),
                )
            )
        return sorted(summaries, key=lambda item: item.clientName)

    def next_month_window(self, today: date | None = None) -> tuple[date, date]:
        today = today or date.today()
        first_next_month = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
        if first_next_month.month == 12:
            following = date(first_next_month.year + 1, 1, 1)
        else:
            following = date(first_next_month.year, first_next_month.month + 1, 1)
        return first_next_month, following - timedelta(days=1)

    def dashboard_payload(
        self,
        office_ids: Iterable[str] | None = None,
        statuses: Iterable[str] | None = None,
        insurers: Iterable[str] | None = None,
        insurance_types: Iterable[str] | None = None,
        expiration_from: date | None = None,
        expiration_to: date | None = None,
    ) -> dict[str, Any]:
        filtered_policies = self.apply_filters(office_ids, statuses, insurers, insurance_types, expiration_from, expiration_to)
        client_summaries = self.build_client_summaries(filtered_policies)
        next_month_start, next_month_end = self.next_month_window()
        expiring_next_month = [
            policy
            for policy in filtered_policies
            if policy.isActive and policy.expirationDate and next_month_start <= policy.expirationDate <= next_month_end
        ]
        cross_sell = [summary for summary in client_summaries if summary.isActiveClient and summary.activePolicyCount == 1]
        former_clients = [summary for summary in client_summaries if summary.isFormerClient]

        return {
            "generatedAt": self.generated_at,
            "sourcePath": self.source_path,
            "filters": {
                "officeIds": list(office_ids or []),
                "statuses": list(statuses or []),
                "insurers": list(insurers or []),
                "insuranceTypes": list(insurance_types or []),
                "expirationFrom": expiration_from.isoformat() if expiration_from else None,
                "expirationTo": expiration_to.isoformat() if expiration_to else None,
            },
            "options": self.options_payload(),
            "metrics": {
                "activeClients": len([summary for summary in client_summaries if summary.isActiveClient]),
                "inactiveClients": len([summary for summary in client_summaries if not summary.isActiveClient]),
                "totalPolicies": len(filtered_policies),
                "singleActivePolicyClients": len(cross_sell),
                "formerClients": len(former_clients),
                "totalPremium": round(sum(policy.premiumNet for policy in filtered_policies), 2),
                "activePolicies": len([policy for policy in filtered_policies if policy.isActive]),
                "insurers": len({policy.insurerName for policy in filtered_policies if policy.insurerName}),
            },
            "highlights": {
                "expiringNextMonth": len(expiring_next_month),
                "formerClients": len(former_clients),
                "crossSell": len(cross_sell),
                "insurerReport": len({policy.insurerName for policy in filtered_policies if policy.insurerName}),
            },
            "insights": self.insights_payload(filtered_policies),
            "charts": self.charts_payload(filtered_policies, client_summaries),
        }

    def charts_payload(self, policies: list[PolicyRecord], client_summaries: list[ClientSummary]) -> dict[str, Any]:
        clients_by_id = {summary.clientId: summary for summary in client_summaries}
        current_year = date.today().year
        monthly_premiums = []
        for month in range(1, 13):
            monthly_policies = [
                policy
                for policy in policies
                if policy.expirationDate and policy.expirationDate.year == current_year and policy.expirationDate.month == month
            ]
            monthly_premiums.append(
                {
                    "month": f"{current_year}-{month:02d}",
                    "label": MONTH_LABELS[month - 1],
                    "premium": round(sum(policy.premiumNet for policy in monthly_policies), 2),
                    "policies": len(monthly_policies),
                }
            )

        insurer_groups: dict[str, list[PolicyRecord]] = defaultdict(list)
        for policy in policies:
            insurer_groups[policy.insurerName or "Sin aseguradora"].append(policy)
        insurer_rows = []
        for insurer, items in insurer_groups.items():
            client_count = len({item.clientId for item in items})
            premium = round(sum(item.premiumNet for item in items), 2)
            insurer_rows.append({"name": insurer, "clients": client_count, "premium": premium})
        insurer_rows.sort(key=lambda row: row["premium"], reverse=True)

        total_clients = max(len(clients_by_id), 1)
        total_premium = sum(max(policy.premiumNet, 0) for policy in policies) or 1
        pie_rows = [
            {
                "name": row["name"],
                "clientsPercent": round((row["clients"] / total_clients) * 100, 2),
                "premiumPercent": round((max(row["premium"], 0) / total_premium) * 100, 2),
            }
            for row in insurer_rows[:8]
        ]

        locality_groups: dict[str, list[PolicyRecord]] = defaultdict(list)
        for policy in policies:
            locality_groups[policy.locality or "Sin localidad"].append(policy)
        total_policy_count = len(policies) or 1
        total_client_count = len({policy.clientId for policy in policies}) or 1
        total_net_premium = sum(max(policy.premiumNet, 0) for policy in policies) or 1
        locality_rows = []
        for locality, locality_policies in locality_groups.items():
            points = [
                point
                for point in (postal_code_point(policy.postalCode) for policy in locality_policies)
                if point
            ]
            if not points:
                continue
            clients = len({policy.clientId for policy in locality_policies})
            policy_count = len(locality_policies)
            premium = round(sum(policy.premiumNet for policy in locality_policies), 2)
            known_coords = locality_coordinates(locality)
            if known_coords:
                lon, lat = known_coords
                min_lon, max_lon = -18.5, 4.5
                min_lat, max_lat = 27.2, 43.9
                x = round(((lon - min_lon) / (max_lon - min_lon)) * 100, 2)
                y = round((1 - ((lat - min_lat) / (max_lat - min_lat))) * 100, 2)
            else:
                x = round(sum(float(point["x"]) for point in points) / len(points), 2)
                y = round(sum(float(point["y"]) for point in points) / len(points), 2)
                lon = round(sum(float(point["lon"]) for point in points) / len(points), 5)
                lat = round(sum(float(point["lat"]) for point in points) / len(points), 5)
            locality_rows.append(
                {
                    "locality": locality,
                    "lon": lon,
                    "lat": lat,
                    "x": x,
                    "y": y,
                    "clients": clients,
                    "policies": policy_count,
                    "premium": premium,
                    "clientsPercent": round((clients / total_client_count) * 100, 2),
                    "policiesPercent": round((policy_count / total_policy_count) * 100, 2),
                    "premiumPercent": round((max(premium, 0) / total_net_premium) * 100, 2),
                }
            )
        locality_rows.sort(key=lambda row: row["policies"], reverse=True)

        return {
            "monthlyPremiums": monthly_premiums,
            "insurerBars": insurer_rows[:10],
            "distribution": pie_rows,
            "localities": locality_rows[:30],
            "localityRows": locality_rows,
        }

    def ranked_policies(self, policies: list[PolicyRecord], key_fn: Any, limit: int | None = None) -> list[dict[str, Any]]:
        grouped: dict[str, list[PolicyRecord]] = defaultdict(list)
        for policy in policies:
            grouped[key_fn(policy)].append(policy)
        rows = [
            {
                "name": name or "Sin dato",
                "policies": len(items),
                "activePolicies": len([item for item in items if item.isActive]),
                "premium": round(sum(item.premiumNet for item in items), 2),
            }
            for name, items in grouped.items()
        ]
        rows.sort(key=lambda row: (row["premium"], row["policies"]), reverse=True)
        return rows[:limit] if limit else rows

    def insights_payload(self, policies: list[PolicyRecord]) -> dict[str, Any]:
        return {
            "byInsurer": self.ranked_policies(policies, lambda policy: policy.insurerName, limit=6),
            "byOffice": self.ranked_policies(policies, lambda policy: self.office_label(policy.officeId), limit=6),
        }

    def expiring_next_month_listing(
        self,
        office_ids: Iterable[str] | None = None,
        statuses: Iterable[str] | None = None,
        insurers: Iterable[str] | None = None,
        insurance_types: Iterable[str] | None = None,
        expiration_from: date | None = None,
        expiration_to: date | None = None,
        today: date | None = None,
    ) -> dict[str, Any]:
        filtered = self.apply_filters(office_ids, statuses, insurers, insurance_types, expiration_from, expiration_to)
        if expiration_from and expiration_to:
            start, end = expiration_from, expiration_to
            title = "Pólizas que vencen en el mes seleccionado"
        else:
            start, end = self.next_month_window(today)
            title = "Pólizas que vencen el próximo mes"
        rows = [
            {
                "clientName": policy.clientName,
                "policyNumber": policy.policyNumber,
                "insuranceType": policy.insuranceType,
                "insurerName": policy.insurerName,
                "email": policy.email,
                "expirationDate": policy.expirationDate.isoformat() if policy.expirationDate else None,
                "officeLabel": self.office_label(policy.officeId),
            }
            for policy in filtered
            if policy.isActive and policy.expirationDate and start <= policy.expirationDate <= end
        ]
        rows.sort(key=lambda row: (row["expirationDate"] or "", row["clientName"], row["policyNumber"]))
        return {
            "title": title,
            "columns": [
                {"key": "clientName", "label": "Cliente"},
                {"key": "policyNumber", "label": "Póliza"},
                {"key": "insurerName", "label": "Aseguradora"},
                {"key": "insuranceType", "label": "Tipo de seguro"},
                {"key": "officeLabel", "label": "Oficina"},
                {"key": "email", "label": "Email"},
                {"key": "expirationDate", "label": "Vencimiento"},
            ],
            "rows": rows,
            "window": {"from": start.isoformat(), "to": end.isoformat()},
        }

    def former_clients_listing(
        self,
        office_ids: Iterable[str] | None = None,
        statuses: Iterable[str] | None = None,
        insurers: Iterable[str] | None = None,
        insurance_types: Iterable[str] | None = None,
        expiration_from: date | None = None,
        expiration_to: date | None = None,
    ) -> dict[str, Any]:
        summaries = self.build_client_summaries(self.apply_filters(office_ids, statuses, insurers, insurance_types, expiration_from, expiration_to))
        rows = [
            {
                "clientName": summary.clientName,
                "email": summary.email,
                "lastPolicyType": summary.lastPolicyType,
                "clientDropDate": summary.clientDropDate.isoformat() if summary.clientDropDate else None,
                "clientDropReason": summary.clientDropReason,
            }
            for summary in summaries
            if summary.isFormerClient
        ]
        rows.sort(key=lambda row: (row["clientName"], row["lastPolicyType"]))
        return {
            "title": "Clientes que fueron clientes y ahora no lo son",
            "columns": [
                {"key": "clientName", "label": "Cliente"},
                {"key": "email", "label": "Email"},
                {"key": "lastPolicyType", "label": "Última póliza"},
                {"key": "clientDropDate", "label": "Fecha de baja"},
                {"key": "clientDropReason", "label": "Motivo de baja"},
            ],
            "rows": rows,
        }

    def cross_sell_listing(
        self,
        office_ids: Iterable[str] | None = None,
        statuses: Iterable[str] | None = None,
        insurers: Iterable[str] | None = None,
        insurance_types: Iterable[str] | None = None,
        expiration_from: date | None = None,
        expiration_to: date | None = None,
    ) -> dict[str, Any]:
        filtered = self.apply_filters(office_ids, statuses, insurers, insurance_types, expiration_from, expiration_to)
        summaries = {summary.clientId: summary for summary in self.build_client_summaries(filtered)}
        rows = []
        for summary in summaries.values():
            if not summary.isActiveClient or summary.activePolicyCount != 1:
                continue
            active_policy = next(
                policy
                for policy in filtered
                if policy.clientId == summary.clientId and policy.isActive
            )
            rows.append(
                {
                    "clientName": summary.clientName,
                    "policyNumber": active_policy.policyNumber,
                    "insurerName": active_policy.insurerName,
                    "insuranceType": active_policy.insuranceType,
                    "email": summary.email,
                    "officeLabel": self.office_label(active_policy.officeId),
                }
            )
        rows.sort(key=lambda row: (row["clientName"], row["insuranceType"]))
        return {
            "title": "Ventas cruzadas",
            "columns": [
                {"key": "clientName", "label": "Cliente"},
                {"key": "policyNumber", "label": "Póliza"},
                {"key": "insurerName", "label": "Aseguradora"},
                {"key": "insuranceType", "label": "Tipo de seguro"},
                {"key": "officeLabel", "label": "Oficina"},
                {"key": "email", "label": "Email"},
            ],
            "rows": rows,
        }

    def insurers_listing(
        self,
        office_ids: Iterable[str] | None = None,
        statuses: Iterable[str] | None = None,
        insurers: Iterable[str] | None = None,
        insurance_types: Iterable[str] | None = None,
        expiration_from: date | None = None,
        expiration_to: date | None = None,
    ) -> dict[str, Any]:
        filtered = self.apply_filters(office_ids, statuses, insurers, insurance_types, expiration_from, expiration_to)
        rows = self.ranked_policies(filtered, lambda policy: policy.insurerName)
        return {
            "title": "Reporte por aseguradora",
            "columns": [
                {"key": "name", "label": "Aseguradora"},
                {"key": "policies", "label": "Pólizas"},
                {"key": "activePolicies", "label": "Pólizas activas"},
                {"key": "premium", "label": "Prima neta"},
            ],
            "rows": rows,
        }

    def _month_window_from_query(self, query: str, today: date | None = None) -> tuple[date, date] | None:
        normalized = normalize_for_match(query)
        today = today or date.today()
        year_match = re.search(r"\b(20\d{2})\b", normalized)
        year = int(year_match.group(1)) if year_match else None
        for month_name, month_number in MONTH_NAMES.items():
            if month_name in normalized:
                resolved_year = year or (today.year + 1 if month_number < today.month else today.year)
                start = date(resolved_year, month_number, 1)
                if month_number == 12:
                    end = date(resolved_year + 1, 1, 1) - timedelta(days=1)
                else:
                    end = date(resolved_year, month_number + 1, 1) - timedelta(days=1)
                return start, end
        if "proximo mes" in normalized or "mes proximo" in normalized or "next month" in normalized:
            return self.next_month_window(today)
        return None

    def _insurance_types_from_query(self, query: str) -> list[str]:
        normalized = normalize_for_match(query)
        matches = set()
        for term, insurance_type in INSURANCE_SYNONYMS.items():
            if term in normalized:
                matches.add(insurance_type)
        for insurance_type in self.insurance_options:
            normalized_type = normalize_for_match(insurance_type)
            if normalized_type and normalized_type in normalized:
                matches.add(insurance_type)
        return sorted(matches)

    def _office_ids_from_query(self, query: str) -> list[str]:
        return sorted(set(re.findall(r"oficina\s+(\d+)", normalize_for_match(query))))

    def chat_response(
        self,
        query: str,
        office_ids: Iterable[str] | None = None,
        statuses: Iterable[str] | None = None,
        insurers: Iterable[str] | None = None,
        insurance_types: Iterable[str] | None = None,
        expiration_from: date | None = None,
        expiration_to: date | None = None,
        today: date | None = None,
    ) -> dict[str, Any]:
        normalized_query = normalize_for_match(query)
        query_offices = self._office_ids_from_query(query)
        query_types = self._insurance_types_from_query(query)
        month_window = self._month_window_from_query(query, today)

        final_offices = sorted({*list(office_ids or []), *query_offices})
        final_statuses = sorted({*list(statuses or [])})
        final_insurers = sorted({*list(insurers or [])})
        final_types = sorted({*list(insurance_types or []), *query_types})
        final_from = expiration_from
        final_to = expiration_to
        if month_window:
            final_from, final_to = month_window

        summary = self.dashboard_payload(final_offices, final_statuses, final_insurers, final_types, final_from, final_to)
        metrics = summary["metrics"]

        if "aseguradora" in normalized_query or "compania" in normalized_query or "compania" in strip_accents(query.lower()):
            listing = self.insurers_listing(final_offices, final_statuses, final_insurers, final_types, final_from, final_to)
            return {
                "answer": f"Encontré {len(listing['rows'])} aseguradoras en el segmento actual.",
                "columns": listing["columns"],
                "rows": listing["rows"][:25],
                "filters": summary["filters"],
            }

        if "ex cliente" in normalized_query or "fueron clientes" in normalized_query or "former client" in normalized_query:
            listing = self.former_clients_listing(final_offices, final_statuses, final_insurers, final_types, final_from, final_to)
            return {
                "answer": f"Encontré {len(listing['rows'])} ex clientes para el criterio actual.",
                "columns": listing["columns"],
                "rows": listing["rows"][:25],
                "filters": summary["filters"],
            }

        if "venta cruzada" in normalized_query or "cross sell" in normalized_query or "una sola poliza" in normalized_query:
            listing = self.cross_sell_listing(final_offices, final_statuses, final_insurers, final_types, final_from, final_to)
            return {
                "answer": f"Hay {len(listing['rows'])} clientes con exactamente una póliza activa en el segmento consultado.",
                "columns": listing["columns"],
                "rows": listing["rows"][:25],
                "filters": summary["filters"],
            }

        if "venc" in normalized_query or "expir" in normalized_query:
            if month_window:
                start, end = month_window
                filtered = self.apply_filters(final_offices, final_statuses, final_insurers, final_types, start, end)
                rows = [
                    {
                        "clientName": policy.clientName,
                        "policyNumber": policy.policyNumber,
                        "insurerName": policy.insurerName,
                        "insuranceType": policy.insuranceType,
                        "email": policy.email,
                        "expirationDate": policy.expirationDate.isoformat() if policy.expirationDate else None,
                    }
                    for policy in filtered
                    if policy.isActive and policy.expirationDate and start <= policy.expirationDate <= end
                ]
                rows.sort(key=lambda row: (row["expirationDate"] or "", row["clientName"]))
                return {
                    "answer": f"Hay {len(rows)} pólizas activas que vencen entre {start.isoformat()} y {end.isoformat()}.",
                    "columns": [
                        {"key": "clientName", "label": "Cliente"},
                        {"key": "policyNumber", "label": "Póliza"},
                        {"key": "insurerName", "label": "Aseguradora"},
                        {"key": "insuranceType", "label": "Tipo de seguro"},
                        {"key": "email", "label": "Email"},
                        {"key": "expirationDate", "label": "Vencimiento"},
                    ],
                    "rows": rows[:25],
                    "filters": summary["filters"],
                }
            listing = self.expiring_next_month_listing(final_offices, final_statuses, final_insurers, final_types, final_from, final_to, today=today)
            return {
                "answer": f"Hay {len(listing['rows'])} pólizas activas que vencen el próximo mes.",
                "columns": listing["columns"],
                "rows": listing["rows"][:25],
                "filters": summary["filters"],
            }

        if "prima" in normalized_query or "monto" in normalized_query:
            return {
                "answer": f"El monto total de primas del conjunto consultado es EUR {metrics['totalPremium']:,.2f}.",
                "columns": [],
                "rows": [],
                "filters": summary["filters"],
            }

        if (
            ("cuanta" in normalized_query or "cuantas" in normalized_query or "cuanto" in normalized_query or "cuantos" in normalized_query or "numero" in normalized_query)
            and ("activo" in normalized_query or "activos" in normalized_query or "vigor" in normalized_query)
            and (query_offices or query_types)
        ):
            return {
                "answer": f"Hay {metrics['activePolicies']} pólizas activas en el segmento consultado.",
                "columns": [],
                "rows": [],
                "filters": summary["filters"],
            }

        if "activos" in normalized_query and "cliente" in normalized_query:
            return {
                "answer": f"Hay {metrics['activeClients']} clientes activos en el segmento actual.",
                "columns": [],
                "rows": [],
                "filters": summary["filters"],
            }

        if ("no activos" in normalized_query or "inactivos" in normalized_query) and "cliente" in normalized_query:
            return {
                "answer": f"Hay {metrics['inactiveClients']} clientes no activos en el segmento actual.",
                "columns": [],
                "rows": [],
                "filters": summary["filters"],
            }

        if "poliza" in normalized_query and ("cuantas" in normalized_query or "cuantos" in normalized_query or "numero" in normalized_query):
            if "activa" in normalized_query or "vigor" in normalized_query:
                answer = f"Hay {metrics['activePolicies']} pólizas activas con los filtros aplicados."
            else:
                answer = f"Hay {metrics['totalPolicies']} pólizas en el conjunto consultado."
            return {"answer": answer, "columns": [], "rows": [], "filters": summary["filters"]}

        return {
            "answer": (
                f"Resumen actual: {metrics['activeClients']} clientes activos, {metrics['inactiveClients']} no activos, "
                f"{metrics['totalPolicies']} pólizas, {metrics['formerClients']} ex clientes y EUR {metrics['totalPremium']:,.2f} en primas."
            ),
            "columns": [],
            "rows": [],
            "filters": summary["filters"],
        }
